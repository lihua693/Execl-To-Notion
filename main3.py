from openpyxl import load_workbook
from notion.client import NotionClient
from notion.block import *
import json, datetime, sys, pandas, getopt
from xls2xlsx import XLS2XLSX


# 获取参数列表
def get_args_func(argv):
    # ret_val = {}
    config = ''  # 默认值
    finish_path = ''
    source_path = ''
    try:
        opts, args = getopt.getopt(argv, "hc:d:s:", ["help", "config=", "finish_path=", "source_path="])
    except getopt.GetoptError:
        print('Error: notion.py -c <config> -d <finish_path> -s <source_path>')
        print('   or: notion.py --config=<config> --finish_path=<finish_path> --source_path=<source_path>')
        sys.exit(2)

    for opt, arg in opts:  # 依次获取列表中的元组项
        if opt in ("-h", "--help"):
            print('Error: notion.py -p <config> -d <finish_path> -s <source_path>')
            print('   or: notion.py --config=<config> --finish_path=<finish_path> --source_path=<source_path>')
            sys.exit()
        elif opt in ("-c", "--config"):
            config = arg
        elif opt in ("-d", "--finish_path"):
            finish_path = arg
        elif opt in ("-s", "--source_path"):
            source_path = arg
    return config, finish_path, source_path


# 连接到notion表位置
def connect_notion(client, url, floor):
    # 连接collection
    page = client.get_block(url)
    collection_ind = floor[-1]
    del (floor[-1])
    target_id = 0
    for i in range(len(floor)):
        target_id = 0
        for j in page.children:
            try:
                if j.title == floor[i]:
                    target_id = j.id
            except Exception as e:
                pass
        if target_id == 0:
            print("请核对floor信息")
            exit(1)
        page = client.get_block(target_id)
    for i in page.children:
        try:
            if i.title == collection_ind:
                target_id = i.collection.id
        except Exception as e:
            pass
    return client.get_collection(target_id)


# 类型转换
def turn_to_notion_type(cell, notion_type):
    if notion_type == "int":
        if isinstance(cell, int) or isinstance(cell, float):
            return cell
        else:
            return float(cell)
    elif notion_type == "date":
        if isinstance(cell, str):
            date_ = datetime.datetime.strptime(cell, '%Y-%m-%d %H:%M:%S')
            return date_
    else:
        return str(cell)


if __name__ == '__main__':
    # 获取断点
    is_error = 0  # 判断是否出错
    with open("./log.json", 'r', encoding="utf-8") as file:
        log = json.load(file)
    if log["ErrorMsg"] is not None:
        is_error = 1
    power_off = log["Power_off"]

    # 获取配置信息
    parameter = get_args_func(sys.argv[1:])
    config = parameter[0]
    finish_path = parameter[1]
    source_path = parameter[2]
    if config == '':
        config = "./package.json"
    with open(config, 'r', encoding="utf-8") as f:
        conf = json.load(f)
    if finish_path == '':
        finish_path = conf["finish"]
    if source_path == '':
        source_path = conf["source"]
    token = conf["token"]
    url = conf["url"]
    info = conf["info"]
    floor = conf["floor"]

    #  判断目录是否存在
    if not os.path.exists(source_path):
        print("数据文件夹不存在！")
        exit(1)
    if not os.path.exists(finish_path) or source_path == finish_path:
        print("目的文件夹不存在！")
        exit(1)
    # 连接Notion
    client = NotionClient(token_v2=token)
    collection = connect_notion(client=client, floor=floor, url=url)
    # 获取数据目录
    files = os.listdir(source_path)
    # 创建工作完成目录
    repeat_file = []
    repeat_line_all = []
    repeat_info_all = []
    for file in files:
        if not file.endswith(".xls") and not file.endswith(".xlsx"):
            continue
        # 读取xlsx
        filename = source_path + "/" + file
        repeat_file.append(file)
        real_path = filename
        if filename.endswith(".xls"):
            x2x = XLS2XLSX(filename)
            x2x.to_xlsx(source_path + "tmp.xlsx")
            filename = + source_path + "tmp.xlsx"

        wb = load_workbook(filename)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        row_num = ws.max_row - 1

        # 获取execl表头
        title = []
        for i in rows:
            for k in i:
                title.append(k)
            break

        # 将迭代对象转换为列表
        datas = []
        for i in rows:
            datas.append(i)

        addRow = 0  # 返回的rowBlock对象，用于插入后删除
        row_add_failed = 0  # 是否插入失败，如果插入失败则删除插入。
        line = 0  # 行位置，结束输出位置信息,用于设置异常退出断点。
        repeat_line = []  # 用于输出第几行什么信息没有添加。
        repeat_info = []
        try:
            print("开始添加：" + file)
            for line in range(len(datas)):
                if is_error == 1 and power_off > line - 1:
                    continue
                print("正在插入：" + file + " ( {} / {} )".format(line + 1, row_num))
                if datas[line] is None:
                    print(file + " ( {} / {} )".format(line + 1, row_num) + "为空行")
                    continue
                add_ind = 0
                # 开始添加行
                add_to_row = {}  # 添加到行的信息
                is_break = 0  # 重复信息停止插入
                for cell in datas[line]:
                    # 将info中execl头与notion表对应
                    execl_title = ""
                    notion_type = ""
                    for i in info:
                        if i["Execl"] == title[add_ind]:
                            execl_title = i["Notion"]
                            notion_type = i["Type"]
                            break
                    # 没用对应上则跳过
                    if execl_title == "":
                        # print("配置文件中的 {} 未匹配成功，请查看配置文件信息。")
                        add_ind += 1
                        continue
                    # 类型匹配
                    if execl_title == conf["primeKey"]:
                        search = client.search(str(cell))
                        if len(search) != 0:
                            repeat_line.append(line + 1)
                            repeat_info.append(cell)
                            print("订单号：{}已存在！".format(cell))
                            is_break = 1
                            break
                    elif cell is None:
                        add_ind += 1
                        continue
                    cell = turn_to_notion_type(cell, notion_type)
                    add_to_row[execl_title] = cell
                    add_ind += 1
                if is_break == 1:
                    continue
                if not add_to_row:
                    print("添加空行，请核对notion与execl对应info")
                    exit(1)
                try:
                    addRow = collection.add_row(**add_to_row)  # 添加行
                except Exception as e:
                    print("error: " + str(e))
                    addRow.remove()
                    with open("./log.json", "w") as f:
                        json.dump(log, f)
                    print("程序异常退出，请调试后重新运行")
                    exit(1)
                # 写入log日志
                log["ErrorMsg"] = "Doing"
                log["Power_off"] = line
                with open("./log.json", "w") as f:
                    json.dump(log, f)
        except Exception as e:
            print("error: " + str(e))
            log["ErrorMsg"] = str(e)
            log["Power_off"] = line - 1
            print(line)
            with open("./log.json", "w") as f:
                json.dump(log, f)
            print("程序异常退出，请调试后重新运行")
            exit(1)
        print(file + "添加完毕！")
        for i in range(len(repeat_line)):
            print("第 {} 行 编号为 {} 插入失败".format(repeat_line[i], repeat_info[i]))
        repeat_info_all.append(repeat_info)
        repeat_line_all.append(repeat_line)
        data = pandas.read_excel(filename)
        data.to_excel(finish_path + "/" + file, index=False)
        os.remove(filename)
        if os.path.exists(real_path):
            os.remove(real_path)

    if len(repeat_file) == 0:
        print("无添加任务！")
    for i in range(len(repeat_file)):
        print(repeat_file[i] + ":")
        if len(repeat_info_all[i]) == 0:
            print("全部添加成功!")
        for j in range(len(repeat_line_all[i])):
            print("第 {} 行 编号为 {} 插入失败".format(repeat_line_all[i][j], repeat_info_all[i][j]))

    log["ErrorMsg"] = None
    log["Power_off"] = 0
    with open("./log.json", "w") as f:
        json.dump(log, f)
    print("恭喜你，成功完成任务！")
